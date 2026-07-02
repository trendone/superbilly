#!/usr/bin/env python3
"""Build import-data.js from the Billy-Liste (sheets Jan 26 .. Dez 26).

Generates a compact ES module exporting IMPORT_DATA = {employees, projects, tasks}
matching the app's data model. Re-run anytime: deterministic output.
"""
import openpyxl, re, json, datetime, sys, os
from collections import Counter, defaultdict

XLSX = os.environ.get("BILLY_XLSX",
    "/Users/petervonaspern/Downloads/BILLY LIST 2016-2023.xlsx")
OUT = os.path.join(os.path.dirname(__file__), "import-data.js")

MONTHS = {'Jan':1,'Feb':2,'Mrz':3,'Mar':3,'Apr':4,'Mai':5,'Jun':6,'Jul':7,
          'Aug':8,'Sep':9,'Okt':10,'Oct':10,'Nov':11,'Dez':12}
COLORS = ['#7c6dfa','#f472b6','#34d399','#fb923c','#60a5fa','#a78bfa',
          '#f87171','#4ade80','#fbbf24','#38bdf8','#e879f9','#a3e635']
VAC, SICK = '__urlaub__', '__krank__'
# absence / non-bookable categories -> (id, display, color). Urlaub/Krank reuse
# the app's system projects (created by ensureSystemProjects), so they are kept
# out of the emitted projects list.
ABS = {
    'urlaub':     (VAC, 'Urlaub', '#64748b'),
    'krank':      (SICK, 'Krank', '#ef4444'),
    'feiertag':   ('feiertag', 'Feiertag', '#94a3b8'),
    'freier tag': ('frei', 'Frei', '#cbd5e1'),
    'frei':       ('frei', 'Frei', '#cbd5e1'),
    'kurzarbeit': ('kurzarbeit', 'Kurzarbeit', '#a8a29e'),
    'uni':        ('uni', 'UNI', '#78716c'),
    'elternzeit': ('elternzeit', 'Elternzeit', '#9ca3af'),
}

def norm(s): return re.sub(r'\s+', ' ', str(s).strip())
def key(s):  return norm(s).lower()
def isnum(v): return isinstance(v, (int, float)) and not isinstance(v, bool)

def daycols(ws):
    # Rohe Tagesspalten (Zeile3=Wochentag, Zeile4=Tag-im-Monat) in Spaltenreihenfolge.
    raw = []
    for c in range(1, ws.max_column + 1):
        dn, dom = ws.cell(3, c).value, ws.cell(4, c).value
        if isinstance(dn, str) and dn.strip().lower() in ('mo','di','mi','do','fr') \
           and isinstance(dom, (int, float)):
            raw.append((c, int(dom)))
    # Diese Kalenderblätter zeigen am Rand die Übertragstage der Nachbarmonate
    # (z.B. das Jul-Blatt beginnt mit 29./30. Juni). Ohne Korrektur würden die auf
    # den Blatt-Monat gestempelt und mit den echten Tagen desselben Datums kollidieren
    # (29. Juni + 29. Juli -> beide 2026-07-29 = Doppelbuchung). Jeder Tag gehört dem
    # Blatt seines echten Monats (existiert separat), daher: Fremdmonats-Spalten weglassen.
    mon = MONTHS.get(ws.title.strip()[:3].title())
    first_one = next((i for i, (c, d) in enumerate(raw) if d == 1), None)
    out, m, prev = [], mon, None
    for i, (c, d) in enumerate(raw):
        if first_one is not None and i < first_one:
            mm = mon - 1 or 12          # führende Übertragstage = Vormonat
        else:
            if prev is not None and d < prev:
                m = m + 1 if m < 12 else 1  # Sprung nach unten = Folgemonat
            mm, prev = m, d
        if mm == mon:                   # nur Tage des eigenen Blatt-Monats behalten
            out.append((c, d))
    return out

def emprows(ws):
    out = []
    for r in range(5, 60):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        a = norm(a)
        if a.upper().startswith('GESAMT') or a.lower().startswith('generelle'):
            break
        out.append((r, a))
    return out

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    sheets = [s for s in wb.sheetnames if re.match(r'^[A-Za-z]{3}\s*26$', s)]

    # Reference roster (row -> name) from the first sheet, to recover names that
    # were overwritten by an absence keyword in the name column (e.g. Okt r14).
    ref = {r: n for r, n in emprows(wb[sheets[0]])}

    # Most common original casing per text key -> canonical display name.
    casing = defaultdict(Counter)
    for s in sheets:
        ws = wb[s]
        for r, _ in emprows(ws):
            for c, _dom in daycols(ws):
                for cc in (c, c + 1):
                    v = ws.cell(r, cc).value
                    if isinstance(v, str) and v.strip():
                        casing[key(v)][norm(v)] += 1

    projects, proj_order = {}, []
    emp, emp_order = {}, []

    def pid_for(text):
        k = key(text)
        if k in ABS:
            pid, nm, col = ABS[k]
            if pid not in projects:
                projects[pid] = {'id': pid, 'name': nm, 'color': col, 'isVacation': True}
                proj_order.append(pid)
            return pid
        if k not in projects:
            disp = casing[k].most_common(1)[0][0] if k in casing else norm(text)
            idx = len([p for p in proj_order if not projects[p].get('isVacation')])
            projects[k] = {'id': 'p%d' % len(proj_order), 'name': disp,
                           'color': COLORS[idx % len(COLORS)], 'budgetDays': None}
            proj_order.append(k)
        return projects[k]['id']

    def eid_for(name):
        k = key(name)
        if k not in emp:
            emp[k] = {'id': 'e%d' % len(emp_order), 'name': norm(name), 'weeklyHours': 40}
            emp_order.append(k)
        return emp[k]['id']

    frag = defaultdict(list)
    for s in sheets:
        ws = wb[s]
        mon = MONTHS.get(s.strip()[:3].title())
        for r, name in emprows(ws):
            # Recover a real person if the name cell holds an absence keyword.
            if key(name) in ABS:
                name = ref.get(r, name)
                if key(name) in ABS:
                    continue  # truly cannot identify -> skip row
            eid = eid_for(name)
            for c, dom in daycols(ws):
                try:
                    d = datetime.date(2026, mon, dom)
                except ValueError:
                    continue
                L, R = ws.cell(r, c).value, ws.cell(r, c + 1).value
                texts = [v for v in (L, R) if isinstance(v, str) and v.strip()]
                nums = [v for v in (L, R) if isnum(v)]
                if not texts:
                    continue
                dk = d.isoformat()
                if len(texts) == 1:
                    b = float(nums[0]) if nums else 1.0
                    b = 1.0 if b <= 0 else min(b, 2.0)
                    frag[eid].append((dk, pid_for(texts[0]), b))
                else:  # split day: two half-day tasks
                    for t in texts:
                        frag[eid].append((dk, pid_for(t), 0.5))

    def nbday(d):
        n = d + datetime.timedelta(days=1)
        while n.weekday() >= 5:
            n += datetime.timedelta(days=1)
        return n

    tasks, tc = [], 0
    for eid, items in frag.items():
        byday = defaultdict(list)
        for dk, pid, b in items:
            byday[dk].append((pid, b))
        seq = []
        for dk in sorted(byday):
            single = len(byday[dk]) == 1
            for pid, b in byday[dk]:
                seq.append((dk, pid, b, single))
        seq.sort(key=lambda x: (x[0], x[1]))
        consumed = set()
        for idx, (dk, pid, b, single) in enumerate(seq):
            if idx in consumed:
                continue
            start = end = dk
            if single:  # merge consecutive business days of the same project+budget
                cur = datetime.date.fromisoformat(dk)
                while True:
                    nb = nbday(cur).isoformat()
                    found = None
                    for j, (dk2, pid2, b2, s2) in enumerate(seq):
                        if j in consumed:
                            continue
                        if dk2 == nb and pid2 == pid and abs(b2 - b) < 1e-9 and s2:
                            found = j
                            break
                    if found is None:
                        break
                    consumed.add(found)
                    end = nb
                    cur = datetime.date.fromisoformat(nb)
            tasks.append({'id': 't%d' % tc, 'projectId': pid, 'employeeId': eid,
                          'startDate': start, 'endDate': end, 'budget': b})
            tc += 1

    out = {
        'employees': [emp[k] for k in emp_order],
        # drop system Urlaub/Krank — the app creates those itself
        'projects': [projects[k] for k in proj_order if projects[k]['id'] not in (VAC, SICK)],
        'tasks': tasks,
    }
    js = json.dumps(out, ensure_ascii=False, separators=(',', ':'))
    header = (
        "/* ===============================================================\n"
        "   Billy-Liste 2026 - Einmal-Import (Mitarbeiter, Projekte, Buchungen)\n"
        "   Generiert aus 'BILLY LIST 2016-2023.xlsx', Blaetter Jan 26-Dez 26.\n"
        "   Quelle: tools/gen_import.py  (deterministisch, jederzeit neu baubar)\n"
        "   Urlaub/Krank referenzieren die System-Projekte der App.\n"
        "   =============================================================== */\n")
    with open(OUT, 'w', encoding='utf-8') as f:
        f.write(header + "export const IMPORT_DATA = " + js + ";\n")

    kb = round(len(open(OUT, 'rb').read()) / 1024, 1)
    print(f"employees={len(out['employees'])} projects={len(out['projects'])} "
          f"tasks={len(out['tasks'])} file={kb}KB")
    print("employees:", ", ".join(e['name'] for e in out['employees']))

if __name__ == '__main__':
    main()
